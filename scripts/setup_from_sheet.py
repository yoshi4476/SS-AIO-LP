# -*- coding: utf-8 -*-
"""ヒアリングシート（Excel）から、サイト一式を作る

これまでは回答を見ながら設定JSONを手で書き、サイトのHTMLも手で用意していた。
転記の間違いに気づけず、記事を作り始めてから作り直しになることがあった。

記入済みのExcelを渡せば、次を一度に作る。
  ・sites/<id>.json          サイト設定
  ・docs/kw-<id>.md          KW計画の置き場
  ・site-<id>/               サイトの骨格（トップ・会社概要・問い合わせ・
                             プライバシー・カテゴリ一覧）
  ・setup-memo-<id>.md       次にやることの控え

使い方:
    python scripts/setup_from_sheet.py <記入済みExcel>            # 確認だけ
    python scripts/setup_from_sheet.py <記入済みExcel> --write    # 実際に作る
"""
import json
import re
import sys
import unicodedata
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent


def cells(ws):
    """「#」列の記号をキーに、記入欄（3列目）を拾う"""
    out = {}
    for row in ws.iter_rows(values_only=True):
        if not row or not row[0]:
            continue
        key = str(row[0]).strip()
        if re.fullmatch(r"[A-K]\d{1,2}", key):
            out[key] = str(row[2]).strip() if len(row) > 2 and row[2] else ""
    return out


def split(text):
    """「A、B / C」のような書き方をまとめて配列にする"""
    if not text:
        return []
    parts = re.split(r"[、,／/\n・]+", text)
    return [p.strip() for p in parts if p.strip()]


def slugify(text, fallback="site"):
    t = unicodedata.normalize("NFKC", text or "").lower()
    t = re.sub(r"[^a-z0-9]+", "-", t).strip("-")
    return t or fallback


def category_rows(ws):
    """カテゴリは表で書かれている。「カテゴリ名（表示用）」の見出しの下を拾う"""
    out, hit = [], False
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        head = str(row[0] or "")
        if "カテゴリ名" in head:
            hit = True
            continue
        if hit:
            if not row[0]:
                if out:
                    break          # 空行が来たら表の終わり
                continue
            if re.fullmatch(r"[A-K]\d{1,2}", str(row[0]).strip()):
                break              # 次の面に入った
            name = str(row[0]).strip()
            slug = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            out.append((name, slug))
    return out


def read(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    a, cats = {}, []
    for ws in wb.worksheets:
        a.update(cells(ws))
        cats += category_rows(ws)
    a["_categories"] = cats
    return a


def to_config(a):
    domain = a.get("A4", "").replace("https://", "").replace("http://", "").strip("/ ")
    # ドメインの先頭だと「media」「www」など、どのクライアントでも同じ名前になる。
    # 意味のある部分を選ぶ。
    parts = [p for p in domain.split(".")
             if p not in ("www", "media", "blog", "lp", "corp", "co", "jp", "com", "net")]
    site_id = slugify(parts[0] if parts else a.get("A3", ""))
    prefix = a.get("A5", "/blog/").strip()
    prefix = "/" + prefix.strip("/").split("/")[0] if prefix.strip("/") else "/blog"

    kind = a.get("A6", "")
    if "Next" in kind:
        site_type = "nextjs-json"
    elif "Markdown" in kind:
        site_type = "external-md"
    elif kind:
        site_type = "external-html"
    else:
        site_type = "self-static"

    cats = {}
    for i, (name, slug) in enumerate(a.get("_categories", []), 1):
        cats[slug or slugify(name, "cat" + str(i))] = name
    if not cats:
        for i, name in enumerate(split(a.get("D1", "")) or ["お役立ち"], 1):
            cats[slugify(name, "cat" + str(i))] = name

    return {
        "id": site_id,
        "name": a.get("A3", "") or "（メディア名）",
        "domain": domain or "example.com",
        "repo": "", "branch": "main",
        "type": site_type,
        "content_dir": "content", "url_prefix": prefix,
        "ga4_property_id": a.get("H1", "") or a.get("K21", ""),
        "kw_plan": "docs/kw-" + site_id + ".md",
        "theme": a.get("B1", ""),
        "audience": a.get("B2", ""),
        # 読者の悩み。Phase 5 のペルソナ採点が「悩みに答えているか」を見る
        "pains": split(a.get("B3", "")),
        "owns": split(a.get("C1", "")),
        "avoid": split(a.get("C2", "")),
        "categories": cats,
        "kw_seeds": {"industries": split(a.get("E1", "")),
                     "intents": split(a.get("E2", ""))},
        "cta_title": a.get("G1", "") or "無料でご相談ください",
        "cta_desc": a.get("G4", ""),
        "cta": {"label": a.get("G3", "") or "無料相談する",
                "url": a.get("G2", ""), "note": a.get("G4", "")},
        "x_tags": [],
        # 会社情報。サイトの雛形とE-E-A-Tの記述に使う。
        # 電話番号は住所と別に持つ。名寄せ（NAP）の検査が住所・電話・社名を
        # 別々に突き合わせるため、1つの欄にまとめると検査が働かない。
        "company": {
            "name": a.get("K14", "") or a.get("A1", ""),
            "corporate_number": a.get("A2", ""),
            "address": a.get("K15", ""),
            "tel": a.get("K16", ""),
            "representative": a.get("K17", ""),
            "founded": a.get("K18", ""),
            "business": a.get("K19", ""),
            "license": a.get("K20", ""),
            "email": a.get("H4", ""),
        },
        "facts": [x for x in (a.get("F1", ""), a.get("F2", ""),
                              a.get("F3", ""), a.get("F4", "")) if x],
    }


def check(cfg):
    ng, warn = [], []
    if not cfg["domain"] or cfg["domain"] == "example.com":
        ng.append("A4 公開ドメインが空です")
    if not cfg["theme"]:
        ng.append("B1 何のメディアかが空です")
    if not cfg["audience"]:
        ng.append("B2 誰に向けたものかが空です")
    if not cfg["owns"]:
        ng.append("C1 扱うテーマの語が空です")
    if not cfg["facts"]:
        ng.append("F 一次情報が空です。ここが無いと引用されない記事になります")
    if not cfg["avoid"]:
        warn.append("C2 扱わない領域が空です。他サイトとの territory 検査が効きません")
    n = len(cfg["kw_seeds"]["industries"]) * len(cfg["kw_seeds"]["intents"])
    if n < 100:
        warn.append("E1×E2 が" + str(n) + "通りしかありません（200通り以上あるとKWが枯れにくい）")
    if not cfg["ga4_property_id"]:
        warn.append("H1 GA4のプロパティIDが空です。流入を測れません")
    if not cfg["company"]["name"]:
        warn.append("運営会社名が空です。E-E-A-Tの評価に影響します")
    return ng, warn


def page(title, body, cfg, _depth=0):
    """サイトの骨格。凝った見た目は後から差し替える前提で、構造だけ整える"""
    c = cfg["company"]
    nav = "\n".join(
        '      <a href="/' + s + '/">' + n + "</a>"
        for s, n in cfg["categories"].items())
    return (
        "<!DOCTYPE html>\n<html lang=\"ja\">\n<head>\n"
        "<meta charset=\"UTF-8\">\n"
        "<meta name=\"viewport\" content=\"width=device-width, initial-scale=1\">\n"
        "<title>" + title + "｜" + cfg["name"] + "</title>\n"
        "<meta name=\"description\" content=\"" + cfg["theme"][:110] + "\">\n"
        "<link rel=\"canonical\" href=\"https://" + cfg["domain"] + "/\">\n"
        # サイトのルートから配信するので絶対パスでよい。「/」と「../」を
        # 混ぜると /../css/ のような、どこも指さないパスになる
        "<link rel=\"stylesheet\" href=\"/css/style.css\">\n"
        "</head>\n<body>\n"
        "<header class=\"site-header\">\n  <div class=\"inner\">\n"
        "    <a class=\"brand\" href=\"/\">" + cfg["name"] + "</a>\n"
        "    <nav class=\"global-nav\">\n" + nav + "\n"
        "      <a href=\"/about/\">運営者情報</a>\n"
        "      <a href=\"/contact/\" class=\"nav-cta\">お問い合わせ</a>\n"
        "    </nav>\n  </div>\n</header>\n"
        "<main class=\"article\">\n" + body + "\n</main>\n"
        "<footer class=\"site-footer\">\n  <div class=\"inner\">\n"
        "    <p class=\"addr\">運営: " + c["name"] + "<br>" + c["address"] + "</p>\n"
        "    <nav>\n"
        "      <a href=\"/about/\">運営者情報</a>\n"
        "      <a href=\"/contact/\">お問い合わせ</a>\n"
        "      <a href=\"/privacy/\">プライバシーポリシー</a>\n"
        "    </nav>\n"
        "    <div class=\"copyright\">© " + c["name"] + "</div>\n"
        "  </div>\n</footer>\n</body>\n</html>\n")


def build_site(cfg, out):
    c = cfg["company"]
    top = (
        "  <header class=\"article-header\">\n"
        "    <h1>" + cfg["name"] + "</h1>\n"
        "    <p class=\"lead\">" + cfg["theme"] + "</p>\n"
        "  </header>\n"
        "  <p>" + cfg["audience"] + "に向けて発信しています。</p>\n"
        "  <section class=\"cta\">\n"
        "    <p class=\"cta-copy\">" + cfg["cta_title"] + "</p>\n"
        "    <a class=\"btn btn-primary\" href=\""
        + (cfg["cta"]["url"] or "/contact/") + "\">" + cfg["cta"]["label"] + "</a>\n"
        "  </section>")

    about = (
        "  <header class=\"article-header\"><h1>運営者情報</h1></header>\n"
        "  <h2>運営会社</h2>\n"
        "  <div class=\"table-wrap\"><table><tbody>\n"
        "    <tr><th>会社名</th><td>" + c["name"] + "</td></tr>\n"
        "    <tr><th>法人番号</th><td>" + c["corporate_number"] + "</td></tr>\n"
        "    <tr><th>代表者</th><td>" + c["representative"] + "</td></tr>\n"
        "    <tr><th>設立</th><td>" + c["founded"] + "</td></tr>\n"
        "    <tr><th>所在地</th><td>" + c["address"] + "</td></tr>\n"
        "    <tr><th>電話番号</th><td>" + c["tel"] + "</td></tr>\n"
        "    <tr><th>事業内容</th><td>" + c["business"] + "</td></tr>\n"
        "    <tr><th>許認可・資格</th><td>" + c["license"] + "</td></tr>\n"
        "  </tbody></table></div>\n"
        "  <h2>編集方針</h2>\n"
        "  <p>実務で得た一次情報をもとに記事を制作しています。事実と意見を分け、"
        "出典を明記します。情報の時点を示し、古くなった内容は定期的に更新します。</p>")

    contact = (
        "  <header class=\"article-header\"><h1>お問い合わせ</h1></header>\n"
        "  <p>" + (cfg["cta_desc"] or "お気軽にご相談ください。") + "</p>\n"
        "  <p>メール: " + c["email"] + "</p>\n"
        "  <!-- フォームは管制塔GASのURLを action に設定して差し替える -->")

    privacy = (
        "  <header class=\"article-header\"><h1>プライバシーポリシー</h1></header>\n"
        "  <p>" + c["name"] + "（以下「当社」）は、個人情報の保護に関する法律および"
        "関連法令を遵守し、以下のとおり個人情報を適切に取り扱います。</p>\n"
        "  <h2>1. 取得する情報</h2>\n"
        "  <p>お問い合わせフォームを通じて、氏名・会社名・メールアドレス・電話番号・"
        "ご相談内容を取得します。</p>\n"
        "  <h2>2. 利用目的</h2>\n"
        "  <p>お問い合わせへの回答、サービスのご案内に利用します。</p>\n"
        "  <h2>3. 第三者提供</h2>\n"
        "  <p>法令に基づく場合を除き、本人の同意なく第三者へ提供しません。</p>\n"
        "  <h2>4. 開示・訂正・削除</h2>\n"
        "  <p>ご本人からのお申し出があった場合、速やかに対応します。</p>\n"
        "  <p>連絡先: " + c["email"] + "</p>")

    # 記事テンプレートのフッターがここへリンクしている。作らないと
    # 公開した全記事がリンク切れになり、ビルドの検査で毎回警告が出る。
    tokusho = (
        "  <header class=\"article-header\">"
        "<h1>特定商取引法に基づく表記</h1></header>\n"
        "  <div class=\"table-wrap\"><table><tbody>\n"
        "    <tr><th>販売事業者</th><td>" + c["name"] + "</td></tr>\n"
        "    <tr><th>代表者</th><td>" + c["representative"] + "</td></tr>\n"
        "    <tr><th>所在地</th><td>" + c["address"] + "</td></tr>\n"
        "    <tr><th>電話番号</th><td>" + c["tel"] + "</td></tr>\n"
        "    <tr><th>メール</th><td>" + c["email"] + "</td></tr>\n"
        "    <tr><th>販売価格</th><td>各サービスのご案内ページに記載します</td></tr>\n"
        "    <tr><th>お支払い方法</th><td>銀行振込</td></tr>\n"
        "    <tr><th>提供時期</th><td>契約後、個別にご案内します</td></tr>\n"
        "    <tr><th>返品・キャンセル</th><td>役務提供の性質上、"
        "提供開始後の返金はいたしかねます</td></tr>\n"
        "  </tbody></table></div>\n"
        "  <p>内容は取り扱うサービスに合わせて必ず書き換えてください。"
        "この雛形のままでは表示義務を満たしません。</p>")

    pages = {
        "index.html": page(cfg["name"], top, cfg),
        "about/index.html": page("運営者情報", about, cfg, 1),
        "contact/index.html": page("お問い合わせ", contact, cfg, 1),
        "privacy/index.html": page("プライバシーポリシー", privacy, cfg, 1),
        "tokushoho/index.html": page("特定商取引法に基づく表記", tokusho, cfg, 1),
    }
    for slug, name in cfg["categories"].items():
        body = ("  <header class=\"article-header\">\n"
                "    <h1>" + name + "</h1>\n"
                "    <p class=\"lead\">" + name + "に関する記事の一覧です。</p>\n"
                "  </header>\n"
                "  <!-- 記事一覧は build.py が生成します -->")
        pages[slug + "/index.html"] = page(name, body, cfg, 1)

    for rel, html in pages.items():
        f = out / rel
        f.parent.mkdir(parents=True, exist_ok=True)
        f.write_text(html, encoding="utf-8", newline="")

    css = ROOT / "site" / "css" / "style.css"
    if css.is_file():
        (out / "css").mkdir(exist_ok=True)
        (out / "css" / "style.css").write_text(
            css.read_text(encoding="utf-8"), encoding="utf-8", newline="")
    return len(pages)


def make_icons(cfg):
    """仮のアプリアイコン。頭文字を1字だけ置いた無地の四角。

    ロゴが届くまでの仮置き。無いと全ページでリンク切れになる。
    """
    made = []
    try:
        from PIL import Image, ImageDraw, ImageFont
    except ImportError:
        return made
    d = ROOT / "site" / "images"
    d.mkdir(parents=True, exist_ok=True)
    ch = (cfg["name"] or "M")[0]
    for size in (180, 192, 512):
        f = d / ("icon-" + str(size) + ".png")
        if f.exists():
            continue
        im = Image.new("RGB", (size, size), "#0b2447")
        dr = ImageDraw.Draw(im)
        font = None
        for name in ("YuGothB.ttc", "meiryob.ttc", "msgothic.ttc"):
            try:
                font = ImageFont.truetype("C:/Windows/Fonts/" + name, int(size * 0.52))
                break
            except OSError:
                continue
        if font is None:
            font = ImageFont.load_default()
        box = dr.textbbox((0, 0), ch, font=font)
        dr.text(((size - box[2] + box[0]) / 2, (size - box[3] + box[1]) / 2),
                ch, font=font, fill="#ffffff")
        im.save(f, optimize=True)
        made.append("site/images/" + f.name)
    return made


def write_support(cfg):
    """設定JSONだけでは動かない。各スクリプトが読む土台をここで揃える。

    無いと落ちる／黙って効かなくなるものを、シートの回答から作る。
    手で作らせると、作り忘れても原因が表に出ないまま数日進んでしまう。
    """
    made = []
    c = cfg["company"]

    # 1. 会社の正規表記。外部掲載・Schema・レポートが全部ここを見る。
    #    表記がぶれると同名の別法人と混ざり、名寄せが壊れる。
    prof = ROOT / "data" / "company_profile.json"
    prof.parent.mkdir(parents=True, exist_ok=True)
    if not prof.exists():
        prof.write_text(json.dumps({
            "_readme": "会社の正規表記（NAP）。外部への掲載・Schema・レポートはすべて"
                       "ここを参照する。表記がぶれると名寄せが壊れ、同名の別法人と"
                       "混ざる。変更するときは掲載済みの記事も直すこと。",
            "name": c["name"], "corporate_number": c["corporate_number"],
            "address": c["address"], "tel": c["tel"], "email": c["email"],
            "founded": c["founded"], "ceo": c["representative"],
            "business": c["business"], "license": c["license"],
            "sites": {cfg["id"]: "https://" + cfg["domain"] + "/"},
        }, ensure_ascii=False, indent=2), encoding="utf-8")
        made.append("data/company_profile.json")

    # 2. 一次情報。記事はここから最低1つ引く。
    #    空のままだと「どのサイトでも書ける記事」になり、AI検索に引用されない。
    facts = ROOT / "data" / "first_party_facts.json"
    if not facts.exists():
        facts.write_text(json.dumps({
            "_readme": "自社でしか出せない一次情報。記事はここから最低1つ引く。"
                       "出典と時点を必ず持たせ、確認できない数値は載せない。",
            "facts": [{"id": cfg["id"] + "-" + str(i), "sites": [cfg["id"]],
                       "topic": [], "text": t, "source": "ヒアリング", "as_of": ""}
                      for i, t in enumerate(cfg["facts"], 1)],
            "pending": [{"note": "数値は出典と集計期間が要る（景品表示法）。"
                                 "確認できるまで記事に書かない", "owner": "クライアント"}],
        }, ensure_ascii=False, indent=2), encoding="utf-8")
        made.append("data/first_party_facts.json")

    # 3. 別サイトへ移すときの置換元
    sc = ROOT / "site.config.json"
    if not sc.exists():
        # ナビはここが元になる。書かないと build.py の既定（別サイトのカテゴリ）が
        # 全ページに出て、存在しないページへのリンクが並ぶ。
        nav = [{"label": name, "url": "/" + slug + "/"}
               for slug, name in cfg["categories"].items()]
        nav.append({"label": cfg["cta"]["label"], "class": "nav-cta",
                    "url": cfg["cta"]["url"] or "/contact/", "cta": "nav_consult"})
        foot = [{"label": "記事一覧", "url": "/blog/"}] + [
            {"label": name, "url": "/" + slug + "/"}
            for slug, name in cfg["categories"].items()] + [
            {"label": "運営者情報", "url": "/about/"},
            {"label": "お問い合わせ", "url": "/contact/"},
            {"label": "プライバシーポリシー", "url": "/privacy/"},
            {"label": "特定商取引法に基づく表記", "url": "/tokushoho/"}]
        sc.write_text(json.dumps({
            "_comment": "このサイトの識別情報とナビゲーション。"
                        "nav / footer_nav は build.py が全ページに出す。",
            "nav": nav, "footer_nav": foot,
            # 記事一覧の下に出る誘導。書かないと別サイトの /lp/ を指したままになる
            "cta_copy": cfg["cta_title"],
            "cta_url": cfg["cta"]["url"] or "/contact/",
            "cta_label": cfg["cta"]["label"],
            "cta_sub": cfg["cta_desc"] or "お気軽にご相談ください",
            "domain": cfg["domain"],
            "site_url": "https://" + cfg["domain"],
            "site_name": cfg["name"],
            "site_tagline": cfg["theme"], "org_name": c["name"],
            "author_name": c["name"] + " 編集部",
            "author_role": cfg["theme"],
            # 記事の署名。実在の個人にする。「編集部」名義はAI検索に信頼されにくく、
            # 引用されにくい。監修者が決まったらここを差し替える。
            "byline": {
                "name": c["representative"] or (c["name"] + " 編集部"),
                "role": (c["name"] + " " + (c["business"] or "")).strip(),
                "bio": " ".join(cfg["facts"][:2]),
                "url": "https://" + cfg["domain"] + "/about/",
            },
            "tel": c["tel"], "address": c["address"],
            "cf_project": cfg["id"], "github_repo": cfg.get("repo", ""),
        }, ensure_ascii=False, indent=2), encoding="utf-8")
        made.append("site.config.json")

    # 4. AI向けのサイト案内。無いと build.py が読み込みで止まる
    llms = ROOT / "site" / "llms.txt"
    llms.parent.mkdir(parents=True, exist_ok=True)
    if not llms.exists():
        llms.write_text(
            "# " + cfg["name"] + "\n> " + (cfg["theme"] or cfg["name"]) + "\n\n"
            "## 運営者情報\n"
            "- " + c["name"] + (("（法人番号 " + c["corporate_number"] + "）")
                                if c["corporate_number"] else "") + "\n"
            + ("- " + c["business"] + "\n" if c["business"] else "")
            + ("- " + c["address"] + "\n" if c["address"] else "")
            + "\n## 主要コンテンツ\n（記事の公開時に自動で追記されます）\n",
            encoding="utf-8")
        made.append("site/llms.txt")

    # 5. このメディア固有の設定。記事を書くときの指示（CLAUDE.md）が
    #    ここを読む。無いと、誰に向けて何を書くかが決まらないまま生成が走る。
    pj = ROOT / "PROJECT.md"
    if not pj.exists():
        cats = "\n".join("| " + n + " | `" + s + "` |"
                         for s, n in cfg["categories"].items())
        pains = "\n".join("- " + x for x in cfg.get("pains", [])) or "- （ヒアリングB3）"
        seeds = cfg["kw_seeds"]
        clusters = "\n".join(
            "- **" + ind + "**: " + " / ".join(ind + " " + it for it in seeds["intents"][:4])
            for ind in seeds["industries"][:6])
        pj.write_text(
            "# PROJECT.md — このメディア固有の設定\n\n"
            "> パイプライン（[CLAUDE.md](CLAUDE.md)）が参照する。\n"
            "> ヒアリングシートから作られた。変えたいときはここを直す。\n\n"
            "## サイト基本情報\n\n"
            "| 項目 | 値 |\n|:--|:--|\n"
            "| メディア名 | " + cfg["name"] + " |\n"
            "| ドメイン | https://" + cfg["domain"] + " |\n"
            "| 記事URL | " + cfg["url_prefix"] + "/{記事スラッグ}/ |\n"
            "| サイト形式 | " + cfg["type"] + " |\n"
            "| メディアの目的 | " + cfg["theme"] + " |\n"
            "| 運営会社 | " + c["name"] + " |\n"
            "| CV | " + cfg["cta_title"] + "（" + (cfg["cta"]["url"] or "/contact/") + "） |\n\n"
            "## E-E-A-T 著者情報\n\n"
            "| 項目 | 値 |\n|:--|:--|\n"
            "| 著者名 | " + c["name"] + " 編集部 |\n"
            "| 監修者 | " + (c["representative"] or "（ヒアリングI3）") + " |\n"
            "| 会社実績 | " + (c["founded"] or "") + " / "
            + (c["license"] or "") + " |\n"
            "| プロフィールURL | https://" + cfg["domain"] + "/about/ |\n\n"
            "## ターゲットペルソナ（Phase 5 ペルソナエージェント用）\n\n"
            "- " + cfg["audience"] + "\n"
            "- 主な悩み:\n" + pains + "\n\n"
            "## 扱う領域 / 扱わない領域\n\n"
            "- 扱う: " + " / ".join(cfg["owns"]) + "\n"
            "- 扱わない: " + " / ".join(cfg["avoid"]) + "\n\n"
            "**扱わない領域の記事は公開の入口で止まる。** 他サイトと同じ話題を"
            "書くと、検索エンジンがどちらを評価するか決められず両方が下がる。\n\n"
            "## 業種別KWクラスター案\n\n" + clusters + "\n\n"
            "## 記事カテゴリ\n\n"
            "| カテゴリ名 | スラッグ |\n|:--|:--|\n" + cats + "\n\n"
            "## CTA設定\n\n"
            "| 項目 | 値 |\n|:--|:--|\n"
            "| 見出し | " + cfg["cta_title"] + " |\n"
            "| ボタン文言 | " + cfg["cta"]["label"] + " |\n"
            "| 誘導先 | " + (cfg["cta"]["url"] or "/contact/") + " |\n"
            "| 応対体制 | " + (cfg["cta_desc"] or "（ヒアリングG4）") + " |\n\n"
            "## 一次情報（全記事に最低1つ入れる）\n\n"
            + "\n".join("- " + x for x in cfg["facts"]) + "\n\n"
            "数値は出典と集計期間を添える（景品表示法）。"
            "確認できない数値は書かない。\n",
            encoding="utf-8")
        made.append("PROJECT.md")

    # 6. 前日までの成功・失敗の記録。翌朝の生成がこれを読んで踏襲する
    kf = ROOT / "kpi_feedback.md"
    if not kf.exists():
        kf.write_text(
            "# KPIフィードバック\n\n"
            "毎日の計測が自動で書き足す。翌朝の記事生成が冒頭で読み込む。\n\n"
            "## 成功パターン（これを踏襲せよ）\n\n（まだありません）\n\n"
            "## 失敗パターン（これを避けよ）\n\n（まだありません）\n\n"
            "## リライト優先度リスト\n\n（まだありません）\n",
            encoding="utf-8")
        made.append("kpi_feedback.md")

    # 7. 業種ごとの記事計画。無いと kw_status が在庫を数えられない
    pillar = ROOT / "docs" / "industry-pillar-plan.md"
    if not pillar.exists():
        # 「**業種**: KW / KW」の形でしか読まれない。表で書くと
        # 在庫が0件と判定され、補充が必要かどうかを見誤る。
        rows = ["# " + cfg["name"] + " 記事計画", "",
                "> 業種 × 切り口で広げる。kw_discover.py がここを起点に補充する。",
                "> 書式を変えると kw_status.py が在庫を数えられなくなる。", ""]
        for ind in cfg["kw_seeds"]["industries"]:
            kws = [ind + " " + it for it in cfg["kw_seeds"]["intents"]]
            rows.append("**" + ind + "**: " + " / ".join(kws))
            rows.append("")
        pillar.write_text("\n".join(rows) + "\n", encoding="utf-8")
        made.append("docs/industry-pillar-plan.md")

    # 8. アプリアイコンとマニフェスト。全ページの<head>から参照されるため、
    #    無いと公開のたびにリンク切れが出て、本物の404が埋もれる。
    man = ROOT / "site" / "manifest.webmanifest"
    if not man.exists():
        man.write_text(json.dumps({
            "name": cfg["name"], "short_name": cfg["name"][:12],
            "description": cfg["theme"][:110], "start_url": "/",
            "display": "standalone", "background_color": "#ffffff",
            "theme_color": "#0b2447",
            "icons": [{"src": "/images/icon-192.png", "sizes": "192x192",
                       "type": "image/png"},
                      {"src": "/images/icon-512.png", "sizes": "512x512",
                       "type": "image/png"}],
        }, ensure_ascii=False, indent=2), encoding="utf-8")
        made.append("site/manifest.webmanifest")
    made += make_icons(cfg)

    # 9. 実行時に書き込む場所。無いまま走ると途中で落ちる
    for d in ("articles", "reports", "data/ranks", "data/ai_citations",
              "data/youtube_transcripts", "site/images"):
        (ROOT / d).mkdir(parents=True, exist_ok=True)
    return made


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if not args:
        raise SystemExit("使い方: python scripts/setup_from_sheet.py <記入済みExcel> [--write]")
    cfg = to_config(read(Path(args[0])))
    ng, warn = check(cfg)

    print("■ " + cfg["name"] + "（" + cfg["id"] + "）")
    print("    ドメイン: " + cfg["domain"] + cfg["url_prefix"] + "/")
    print("    種別: " + cfg["type"])
    print("    カテゴリ: " + " / ".join(cfg["categories"].values()))
    ind = len(cfg["kw_seeds"]["industries"])
    itn = len(cfg["kw_seeds"]["intents"])
    print("    KWの起点: 業種" + str(ind) + " × 意図" + str(itn)
          + " = " + str(ind * itn) + "通り")
    print("    一次情報: " + str(len(cfg["facts"])) + "件")
    for m in ng:
        print("    × " + m)
    for m in warn:
        print("    ! " + m)
    if ng:
        raise SystemExit("\n  記入漏れがあるため作成しません。埋めてから再実行してください")
    if "--write" not in sys.argv:
        print("\n  確認のみ（--write を付けると作成します）")
        return

    (ROOT / "sites").mkdir(exist_ok=True)
    (ROOT / "sites" / (cfg["id"] + ".json")).write_text(
        json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")
    plan = ROOT / cfg["kw_plan"]
    plan.parent.mkdir(parents=True, exist_ok=True)
    plan.write_text("# " + cfg["name"] + " KW計画\n\n対象: " + cfg["audience"]
                    + "\n\n（kw_discover.py が自動補充します）\n", encoding="utf-8")
    # 当社が作るサイトは site/ がビルドの対象。別フォルダに作ると
    # build.py が見に行かず、ページが1枚も出ないまま気づけない。
    # 先方のサイトへ記事だけ納める場合は、渡す用に site-<id>/ へ出す。
    own = cfg["type"] == "self-static"
    out = ROOT / "site" if own else ROOT / ("site-" + cfg["id"])
    n_page = build_site(cfg, out)
    # 先方のサイトへ納める場合も、手元の site/ に固定ページが要る。
    # 記事テンプレートのフッターが /about/ などを指しているため、
    # 無いとビルドのリンク検査が毎回警告を出し、本物の404に気づけなくなる。
    if not own:
        build_site(cfg, ROOT / "site")
    made = write_support(cfg)

    memo = ROOT / ("setup-memo-" + cfg["id"] + ".md")
    memo.write_text(
        "# " + cfg["name"] + " セットアップの控え\n\n"
        "## 作られたもの\n"
        "- sites/" + cfg["id"] + ".json\n"
        "- " + cfg["kw_plan"] + "\n"
        "- " + out.name + "/（" + str(n_page) + "ページ）\n"
        + "".join("- " + m + "\n" for m in made) + "\n"
        "## 次にやること\n"
        "1. サイトを公開先へ設置し、repo を sites/" + cfg["id"] + ".json に書く\n"
        "2. GA4とSearch Consoleにサービスアカウントを追加（Search Consoleはオーナー）\n"
        "3. python scripts/kw_discover.py --site " + cfg["id"] + " --deep --append\n"
        "4. python scripts/kw_status.py で未着手が60件以上あるか確認\n"
        "5. python scripts/publish_flow.py " + cfg["id"] + " <slug> で1本目\n\n"
        "## 一次情報（記事に必ず入れる）\n"
        + "\n".join("- " + x for x in cfg["facts"]) + "\n", encoding="utf-8")

    print("\n  作成: sites/" + cfg["id"] + ".json")
    print("  作成: " + cfg["kw_plan"])
    print("  作成: " + out.name + "/（" + str(n_page) + "ページ）")
    for m in made:
        print("  作成: " + m)
    print("  作成: " + memo.name)


if __name__ == "__main__":
    main()
